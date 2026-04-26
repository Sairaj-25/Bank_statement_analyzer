import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
from app.models.schemas import AnalysisResult
from app.utils.logger import get_logger

logger = get_logger(__name__)

class ExcelExporter:
    @staticmethod
    def _format_sheet(ws):
        """Applies professional styling to the Excel sheet"""
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def export(df: pd.DataFrame, analysis: AnalysisResult, file_path: str):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Sheet 1: Transactions
            df.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Sheet 2: Monthly Summary (Include Trend %)
            monthly_df = pd.DataFrame([m.model_dump() for m in analysis.monthly_summary])
            monthly_df.to_excel(writer, sheet_name='Monthly Summary', index=False)
            
            # Sheet 3: Category Summary
            cat_df = pd.DataFrame([c.model_dump() for c in analysis.category_summary])
            cat_df.to_excel(writer, sheet_name='Category Summary', index=False)

        # --- POST-PROCESSING FOR STYLING & CHARTS ---
        wb = load_workbook(file_path)
        
        # 1. Format all sheets
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            ExcelExporter._format_sheet(ws)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20

        # 2. Add Pie Chart for Categories
        ws_cat = wb['Category Summary']
        pie = PieChart()
        pie.title = "Spending by Category"
        data = Reference(ws_cat, min_col=2, min_row=1, max_row=ws_cat.max_row)
        cats = Reference(ws_cat, min_col=1, min_row=2, max_row=ws_cat.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        ws_cat.add_chart(pie, "E2")

        # 3. Add Bar Chart for Top Merchants
        if analysis.top_merchants:
            ws_merch = wb.create_sheet("Top Merchants")
            merchant_df = pd.DataFrame(analysis.top_merchants)

            # Write header row manually — ws_merch is an openpyxl Worksheet,
            # not a file path, so pandas .to_excel() cannot be used here.
            headers = list(merchant_df.columns)
            ws_merch.append(headers)
            for record in merchant_df.itertuples(index=False):
                ws_merch.append(list(record))

            ExcelExporter._format_sheet(ws_merch)
            for col in ws_merch.columns:
                ws_merch.column_dimensions[col[0].column_letter].width = 20

            max_row = len(analysis.top_merchants) + 1  # +1 for header
            bar = BarChart()
            bar.type = "bar"
            bar.title = "Top 5 Merchants by Spend"
            data = Reference(ws_merch, min_col=2, min_row=1, max_row=max_row)
            cats = Reference(ws_merch, min_col=1, min_row=2, max_row=max_row)
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            ws_merch.add_chart(bar, "E2")

        wb.save(file_path)
        logger.info(f"Upgraded Excel report generated at {file_path}")